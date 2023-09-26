import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from time import sleep


numero_oab = 133864

driver = webdriver.Chrome()

driver.get('https://pje-consulta-publica.tjmg.jus.br/')
sleep(3)

# Searching the oab number
oab_field = driver.find_element(By.XPATH, '/html/body/div[5]/div/div/div/div[2]/form/div[1]/div/div/div/div/div[9]/div/div[2]/input[1]')
oab_field.send_keys(numero_oab)

dropdown_states = driver.find_element(By.XPATH, '/html/body/div[5]/div/div/div/div[2]/form/div[1]/div/div/div/div/div[9]/div/div[2]/select')
states_options = Select(dropdown_states)
states_options.select_by_visible_text('SP')
sleep(1)

search_button = driver.find_element(By.XPATH, '/html/body/div[5]/div/div/div/div[2]/form/div[1]/div/div/div/div/div[10]/div/input')
search_button.click()
sleep(5)

# Opening the proccess
#proccess = driver.find_elements(By.XPATH, '//*[@id="fPP:processosTable:600834760:j_id243"]/a')
proccess = driver.find_elements(By.XPATH, "//b[@class='btn-block']")
for processo in proccess:
    processo.click()
    sleep(5)
    windows = driver.window_handles
    driver.switch_to.window(windows[-1]) # changing google windows
    driver.set_window_size(1920, 1080)
    #proccess_number = driver.find_elements(By.XPATH, '//*[@id="j_id132:processoTrfViewView:j_id138"]/div/div[2]/div')
    proccess_number = driver.find_elements(By.XPATH, "//div[@class='col-sm-12 ']")
    proccess_number = proccess_number[0]
    proccess_number = proccess_number.text

    sleep(2)

    #data_distribuition = driver.find_elements(By.XPATH, '//*[@id="j_id132:processoTrfViewView:j_id150"]/div/div[2]')
    data_distribuition = driver.find_elements(By.XPATH, "//div[@class='value col-sm-12 ']")
    data_distribuition = data_distribuition[1]
    data_distribuition = data_distribuition.text
    
    movement = driver.find_elements(By.XPATH, "//div[@id='j_id132:processoEventoPanel_body']//tr[contains(@class, 'rich-table-row')]//td//div//div//span")
    movement_list = []
    for move in movement:
        movement_list.append(move.text)

    workbook = openpyxl.load_workbook('dados.xlsx')
    try:
        # código para inserir dados em página já existente
        proccess_page = workbook[proccess_number]
        # criar nome das colunas
        proccess_page['A1'].value = "Número do Processo"
        proccess_page['B1'].value = "Data de distribuição"
        proccess_page['C1'].value = "Movimentações"
        # adicionar número de processo
        proccess_page['A2'].value = proccess_number
        proccess_page['B2'].value = data_distribuition
        # adicionar movimentações
        for index, line in enumerate(proccess_page.iter_rows(min_row=2,max_row=len(movement_list),min_col=3,max_col=3)):
            for celula in line:
                celula.value = movement_list[index]
        workbook.save('dados.xlsx')
        driver.close()
        sleep(3)
        driver.switch_to.window(driver.window_handles[0])
    except Exception as error:
        # código para criar uma página do zero e inserir as informações
        workbook.create_sheet[proccess_number]
        # acessar página do processo
        proccess_page = workbook[proccess_number]
        # criar nome das colunas
        proccess_page['A1'].value = "Número do Processo"
        proccess_page['B1'].value = "Data de distribuição"
        proccess_page['C1'].value = "Movimentações"
        # adicionar número de processo
        proccess_page['A2'].value = proccess_number
        proccess_page['B2'].value = data_distribuition
        # adicionar movimentações
        for index, line in enumerate(proccess_page.iter_rows(min_row=2,max_row=len(movement_list),min_col=3,max_col=3)):
            for celula in line:
                celula.value = movement_list[index]
        workbook.save('dados.xlsx')
        driver.close()
        driver.switch_to.window(driver.window_handles[0])